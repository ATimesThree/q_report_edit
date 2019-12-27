import openpyxl
from openpyxl import load_workbook
from datetime import date
import getpass

print("Working.. Please wait..")

original_file_name = "FILENAME.xlsx"
new_file_name= "REPORT_NAME " + str(date.today()) + ".xlsx"

path_open = "C:/Users/" + getpass.getuser() + "/Desktop/" + original_file_name
path_save = "C:/Users/" + getpass.getuser() + "/Desktop/" + new_file_name

wb = openpyxl.load_workbook(path_open) 
sheet=wb["SHEET_NAME"]
sheet.title = "NEW_SHEET_TITLE"


print("Deleting columns..")

sheet.delete_cols(4,2)
sheet.delete_cols(5,2)
sheet.delete_cols(8,4)


print("Renaming columns..")

sheet["A1"].value = "NEW_NAME_1"
sheet["B1"].value = "NEW_NAME_2"
sheet["C1"].value = "NEW_NAME_3"
sheet["D1"].value = "NEW_NAME_4"
sheet["E1"].value = "NEW_NAME_5"
sheet["F1"].value = "NEW_NAME_6"
sheet["G1"].value = "NEW_NAME_7"


print("Finding DESIRED INFO..")

i=2
while i <= sheet.max_row: 
	
	ref_cell = str(sheet.cell(row=i, column=6).value) 
	if (ref_cell).startswith("START_1") or (ref_cell).startswith("START_2") or (ref_cell).startswith("START_3"):
		i += 1
	else:
		sheet.delete_rows(i,1)


print("Translating the status..")

j = 2
while j <= sheet.max_row:
	status_cell = str(sheet.cell(row=j, column=4).value)
	if status_cell == "STATUS_1" or status_cell == "STATUS_2":
		sheet["D"+str(j)] = "STATUS_3"
		j += 1
	else:
		sheet["D"+str(j)] = "STATUS_4"
		j += 1


print("Saving..")

wb.save(path_save)

print("Saved as: " + new_file_name)